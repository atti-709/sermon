# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Write `_SPEAKERS.txt` into every CampFest sermon folder, off the programme.

The web UI pre-fills a clip's intro card with the first name in the sermon folder's
`_SPEAKERS.txt` (see `sermon.web.projects.load_speakers`). This fills those files in
from the festival programme spreadsheet, so nobody has to retype "MUDr. Janka
Nosková, MPH" from a run sheet at 1 a.m. with the diacritics right.

    uv run scripts/speaker_names.py \\
        --schedule "~/Movies/Campfest2026/SCHEDULE/CF 2026 - PROGRAM.xlsx" \\
        ~/Movies/Campfest2026 ~/Downloads/CAMPFEST26_WIP/STAFF/CAMPFEST26

A root is any tree holding `CF/<day>/VIDEOS/04_SERMONS/<slug>/` folders — the
layout `make_nas_tree.sh` builds, one folder per preached block. Folders are never
created and an existing file is never overwritten without `--force`: the names in
it are meant to be corrected by hand, and re-running must not undo that.
"""

import argparse
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

SERMONS_DIR = "VIDEOS/04_SERMONS"
SPEAKERS_FILE = "_SPEAKERS.txt"

# The programme rows that are somebody preaching, keyed by how `popis` starts, and
# the sermon folder each one belongs to. Longest prefix first, so "recap téma" is
# not read as the morning "téma".
#
# `Seminár` maps to no fixed folder: a seminar's folder is named after whoever gives
# it (`seminar-livia-halmkan`), so those are matched by name — see `folder_for`. The
# other three are named after the block they sit in, which no cell of the programme
# says, so the mapping has to be written down here.
BLOCK_FOLDERS = (
    ("recap téma", "vecerny-program"),
    ("evanjel", "vecerny-program"),
    ("seminár", None),
    ("téma", "rano-worship-tema"),
)

# Slovak letters NFKD leaves alone; same table `extract_sessions.py` slugs with, so
# the slugs computed here match the folder names already on disk.
EXTRA_MAP = str.maketrans({"ľ": "l", "ď": "d", "ť": "t", "ň": "n", "ĺ": "l", "ŕ": "r", "ů": "u"})

# words that belong to a person's name without being capitalised
NAME_PARTICLES = {"a", "&", "st.", "ml.", "s", "so", "and", "jr.", "sr."}

SEPARATORS = (" - ", " – ", " — ", ": ")


def slug(text: str) -> str:
    """ASCII lowercase-hyphen slug, the same rule the folder names were built with."""
    text = text.translate(EXTRA_MAP)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.encode("ascii", "ignore").decode("ascii").lower()
    text = re.sub(r"[\[\]()]", "", text)  # 'Na pravde (ne)záleží' is one word, not three
    return re.sub(r"[^a-z0-9]+", "-", text).strip("-")


def looks_like_a_name(text: str) -> bool:
    """Is this a person, rather than the title of their talk?

    Every word capitalised (bar the particles a name may contain), no more than
    five of them, no sentence punctuation. 'MUDr. Janka Nosková, MPH' passes;
    'Tvoje Slovo je pravda' does not, on 'je'."""
    words = text.split()
    if not 1 <= len(words) <= 5 or any(c in text for c in "?!…"):
        return False
    return all(w[0].isupper() or w.lower() in NAME_PARTICLES for w in words)


def speaker_of(performer: str) -> str:
    """The speaker out of a programme cell: 'Slavo Poloha - Pravdovravnosť' -> 'Slavo Poloha'.

    Almost every cell reads `<name> - <title of the talk>`, but a few are the other
    way round ('Tvoje Slovo je pravda - Tomáš Štrba'), so which side is the person
    is decided by looking at both. Neither side looking like a name means the cell
    is worded some third way — take the left, which is right far more often."""
    for sep in SEPARATORS:
        if sep in performer:
            left, right = (part.strip() for part in performer.split(sep, 1))
            if looks_like_a_name(left) or not looks_like_a_name(right):
                return left
            return right
    return performer.strip()


# --- reading the programme ---------------------------------------------------


def sheet_day(title: str) -> str | None:
    """'štvrtok 6.8.2026' -> '6.8.', the folder name for that day."""
    match = re.search(r"(\d{1,2})\.(\d{1,2})\.", title)
    return f"{int(match.group(1))}.{int(match.group(2))}." if match else None


def cell_text(value: object) -> str:
    return " ".join(str(value).split()) if value is not None else ""


def venue_columns(sheet) -> dict[int, str]:
    """Which venue each `OD` column group belongs to.

    A day's sheet puts the venues side by side — HANGÁR at A, HUMNO at H, STODOLA
    at O … — each a group of OD | DO | dĺžka | popis | účinkujúci under its name in
    row 1, with narrow 'Rekvizity' columns wedged between them."""
    venues = {
        cell.column: cell_text(cell.value)
        for cell in sheet[1]
        if cell_text(cell.value) and cell_text(cell.value).lower() != "rekvizity"
    }
    groups = {}
    for row in sheet.iter_rows(min_row=1, max_row=8):
        for cell in row:
            if cell_text(cell.value).upper() == "OD":
                start = max((c for c in venues if c <= cell.column), default=None)
                groups[cell.column] = venues.get(start, "?")
        if groups:
            break
    return groups


class Block:
    """One preached slot on the programme."""

    def __init__(self, day: str, weekday: str, venue: str, start: str, popis: str, performer: str):
        self.day, self.weekday, self.venue = day, weekday, venue
        self.start, self.popis, self.performer = start, popis, performer
        self.speaker = speaker_of(performer)

    def __str__(self) -> str:
        return f"{self.start:>5}  {self.popis}  ·  {self.performer}"


def preached_blocks(schedule: Path) -> list[Block]:
    """Every row of the programme that is somebody preaching, in programme order."""
    book = openpyxl.load_workbook(schedule, data_only=True)
    blocks = []
    for sheet in book.worksheets:
        day = sheet_day(sheet.title)
        if day is None:
            continue
        weekday = sheet.title.split()[0]
        columns = venue_columns(sheet)
        for row in sheet.iter_rows(min_row=2):
            cells = {cell.column: cell for cell in row}
            for column, venue in columns.items():
                popis = cell_text(cells[column + 3].value) if column + 3 in cells else ""
                performer = cell_text(cells[column + 4].value) if column + 4 in cells else ""
                if not performer or not any(
                    popis.lower().startswith(prefix) for prefix, _ in BLOCK_FOLDERS
                ):
                    continue
                start = cell_text(cells[column].value if column in cells else "")[:5]
                blocks.append(Block(day, weekday, venue, start, popis, performer))
    return blocks


def folder_for(block: Block, folders: dict[str, Path]) -> Path | None:
    """The sermon folder this block's recording goes in, if that folder exists.

    A seminar's folder is the head of the slug of its whole `účinkujúci` cell —
    that is how `make_nas_tree.sh` named them, truncated at a word boundary — so
    `Lívia Halmkan - Pornografia naša každodenná…` finds `seminar-livia-halmkan`.
    Every other kind of block has its folder named in `BLOCK_FOLDERS`."""
    for prefix, folder in BLOCK_FOLDERS:
        if not block.popis.lower().startswith(prefix):
            continue
        if folder is not None:
            return folders.get(folder)
        wanted = slug(block.performer)
        matches = [
            path
            for name, path in folders.items()
            if name.startswith("seminar-") and wanted.startswith(name.removeprefix("seminar-"))
        ]
        return matches[0] if len(matches) == 1 else None
    return None


# --- writing the files -------------------------------------------------------


def render(blocks: list[Block], schedule: Path) -> str:
    """The `_SPEAKERS.txt` for one sermon folder.

    Every name is followed by the programme rows it came from, as comments: the
    person correcting this file is looking at a run sheet, and needs to see which
    slot each name is for before trusting it."""
    first = blocks[0]
    lines = [
        f"# Speaker names for this sermon — {first.weekday} {first.day} · {first.venue}",
        f"# Taken from {schedule.name} by scripts/speaker_names.py. Correct anything wrong.",
        "#",
        *(f"#   {block}" for block in blocks),
        "#",
        "# One name per line, in programme order. The first one pre-fills the intro card",
        "# for every clip cut from this sermon; the others are offered as buttons beside",
        "# the field. Editing the name in the app only changes that one clip and never",
        "# writes back here, so this file stays the programme's version.",
        "",
    ]
    seen = []
    for block in blocks:
        if block.speaker not in seen:
            seen.append(block.speaker)
            lines.append(block.speaker)
    return "\n".join(lines) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("roots", nargs="+", type=Path, help="trees holding CF/<day>/VIDEOS/04_SERMONS")
    parser.add_argument("--schedule", required=True, type=Path, help="the programme .xlsx")
    parser.add_argument("--force", action="store_true", help="overwrite files that already exist")
    parser.add_argument("--dry-run", action="store_true", help="print what would be written")
    args = parser.parse_args()

    schedule = args.schedule.expanduser()
    if not schedule.is_file():
        print(f"no such schedule: {schedule}", file=sys.stderr)
        return 2

    blocks = preached_blocks(schedule)
    print(f"{schedule.name}: {len(blocks)} preached blocks\n")

    written = skipped = 0
    for root in args.roots:
        root = root.expanduser()
        for sermons in sorted(root.glob(f"CF/*/{SERMONS_DIR}")):
            day = sermons.parents[1].name
            folders = {p.name: p for p in sorted(sermons.iterdir()) if p.is_dir()}
            grouped: dict[Path, list[Block]] = {}
            for block in (b for b in blocks if b.day == day):
                folder = folder_for(block, folders)
                if folder is None:
                    continue  # a preached block with no folder — a discussion, a talk we don't cut
                grouped.setdefault(folder, []).append(block)

            for folder in folders.values():
                target = folder / SPEAKERS_FILE
                if folder not in grouped:
                    print(f"  ?  {folder.relative_to(root)} — no block on the programme matched")
                    continue
                if target.is_file() and not args.force:
                    skipped += 1
                    print(f"  ·  {folder.relative_to(root)} — exists, kept (--force to replace)")
                    continue
                names = ", ".join(b.speaker for b in grouped[folder])
                print(f"  ✓  {folder.relative_to(root)} — {names}")
                if not args.dry_run:
                    target.write_text(render(grouped[folder], schedule), encoding="utf-8")
                    written += 1

    verb = "would write" if args.dry_run else "wrote"
    print(f"\n{verb} {written} file(s), kept {skipped}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
